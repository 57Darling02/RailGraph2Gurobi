from __future__ import annotations

from datetime import datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional

from core.types import (
    AnalyzeConfig,
    AppConfig,
    BuildConfig,
    DelayScenario,
    ExportTimetableConfig,
    InputConfig,
    InterruptionScenario,
    ProjectConfig,
    RawTable,
    ScenarioConfig,
    SolveConfig,
    SolverConfig,
    SpeedLimitScenario,
)

TIME_HEADERS = {"arrival_time", "departure_time"}


def _require_yaml() -> Any:
    try:
        import yaml
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Missing dependency: pyyaml") from exc
    return yaml


def _normalize_header(header: str) -> str:
    return header.lower()


def _normalize_time_text(value: str) -> str:
    parts = value.split(":")
    if len(parts) != 3 or any(part == "" for part in parts):
        return value
    try:
        hour, minute, second = [int(part) for part in parts]
    except ValueError:
        return value
    if hour < 0 or hour > 23 or minute < 0 or minute > 59 or second < 0 or second > 59:
        return value
    return f"{hour:02d}:{minute:02d}:{second:02d}"


def _normalize_cell_value(cell: Any, header: str) -> Optional[str]:
    if cell is None:
        return None
    if isinstance(cell, datetime):
        cell = cell.time()
    if isinstance(cell, time):
        return cell.strftime("%H:%M:%S")
    text = str(cell).strip()
    if text == "":
        return None
    if header in TIME_HEADERS:
        return _normalize_time_text(text)
    return text


def _read_excel(path: Path, sheet_name: str) -> RawTable:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Missing dependency: openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Empty sheet: {path}#{sheet_name}")

    raw_headers = [_normalize_header(str(cell).strip()) if cell is not None else "" for cell in rows[0]]
    headers: List[str] = []
    header_indexes: List[int] = []
    for index, header in enumerate(raw_headers):
        if header == "":
            continue
        if header in headers:
            raise ValueError(f"Duplicated header in {path}#{sheet_name}: {header}")
        headers.append(header)
        header_indexes.append(index)

    records: List[Dict[str, Optional[str]]] = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        record: Dict[str, Optional[str]] = {}
        for header, index in zip(headers, header_indexes):
            cell = row[index] if index < len(row) else None
            record[header] = _normalize_cell_value(cell, header)
        records.append(record)
    return RawTable(headers=headers, rows=records)


def _parse_time_to_seconds(value: Any) -> int:
    # YAML may parse unquoted "HH:MM:SS" into a numeric scalar (e.g. 08:00:00 -> 28800).
    # Support both explicit "HH:MM:SS" strings and numeric seconds.
    if isinstance(value, (int, float)):
        seconds = int(value)
        if seconds < 0 or seconds > 24 * 3600 - 1:
            raise ValueError(f"Invalid HH:MM:SS time: {value}")
        return seconds

    text = str(value).strip()
    if text.isdigit():
        seconds = int(text)
        if seconds < 0 or seconds > 24 * 3600 - 1:
            raise ValueError(f"Invalid HH:MM:SS time: {value}")
        return seconds

    parts = text.split(":")
    if len(parts) != 3:
        raise ValueError(f"Invalid HH:MM:SS time: {value}")
    hour, minute, second = [int(part) for part in parts]
    if hour < 0 or hour > 23 or minute < 0 or minute > 59 or second < 0 or second > 59:
        raise ValueError(f"Invalid HH:MM:SS time: {value}")
    return hour * 3600 + minute * 60 + second


def _path_or_default(value: Any, default: Path) -> Path:
    if value is None:
        return default
    text = str(value).strip()
    if text == "":
        return default
    return Path(text)


def _str_or_default(value: Any, default: str) -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text or default


def _bool_or_default(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _required_path(value: Any, field_name: str) -> Path:
    if value is None:
        raise ValueError(f"Missing required config field: {field_name}")
    text = str(value).strip()
    if text == "":
        raise ValueError(f"Missing required config field: {field_name}")
    return Path(text)


def load_config(path: Path) -> AppConfig:
    yaml = _require_yaml()
    with path.open("r", encoding="utf-8") as file:
        payload = yaml.safe_load(file) or {}

    project_cfg = payload.get("project", {}) or {}
    build_cfg = payload.get("build", {}) or {}
    solve_cfg = payload.get("solve", {}) or {}
    export_cfg = payload.get("export-timetable", payload.get("export_timetable", {})) or {}
    analyze_cfg = payload.get("analyze", payload.get("analysis", {})) or {}

    # Backward-compatible sections.
    input_cfg = payload.get("input", {}) or {}
    root_solver_cfg = payload.get("solver", {}) or {}
    root_scenarios_cfg = payload.get("scenarios", {}) or {}

    case_name = _str_or_default(project_cfg.get("name"), path.stem)
    output_dir = _path_or_default(project_cfg.get("output_dir"), Path("outputs") / case_name)

    # Convention-first artifact paths.
    # build output: output_dir/<name>.lp
    # solve output: output_dir/<name>.sol
    # export output: output_dir/adjusted_timetable.xlsx
    lp_default_path = output_dir / f"{case_name}.lp"
    solution_default_path = output_dir / f"{case_name}.sol"
    adjusted_default_path = output_dir / "adjusted_timetable.xlsx"
    metrics_default_path = output_dir / "analysis_metrics.xlsx"
    plot_default_path = output_dir / "timetable_plot.png"

    # project holds base input info in new schema; fall back to legacy input section.
    timetable_path = _required_path(
        project_cfg.get("timetable_path", input_cfg.get("timetable_path")),
        "project.timetable_path (or legacy input.timetable_path)",
    )
    mileage_path = _required_path(
        project_cfg.get("mileage_path", input_cfg.get("mileage_path")),
        "project.mileage_path (or legacy input.mileage_path)",
    )
    timetable_sheet_name = _str_or_default(
        project_cfg.get("timetable_sheet_name", input_cfg.get("timetable_sheet_name")),
        "Sheet1",
    )
    mileage_sheet_name = _str_or_default(
        project_cfg.get("mileage_sheet_name", input_cfg.get("mileage_sheet_name")),
        "Sheet1",
    )

    # build reads build.scenarios first, then fallback to legacy top-level scenarios.
    scenarios_cfg = build_cfg.get("scenarios", root_scenarios_cfg) or {}

    solve_solver_cfg = solve_cfg.get("solver", {}) or {}

    def _solve_value(key: str, default: Any) -> Any:
        if key in solve_cfg:
            return solve_cfg[key]
        if key in solve_solver_cfg:
            return solve_solver_cfg[key]
        return root_solver_cfg.get(key, default)

    solve_lp_path = _path_or_default(solve_cfg.get("lp_path"), lp_default_path)
    export_solution_path = _path_or_default(
        export_cfg.get("sol_path", export_cfg.get("solution_path")),
        solution_default_path,
    )

    adjusted_timetable_path = _path_or_default(
        analyze_cfg.get("adj_timetable_path", analyze_cfg.get("adjusted_timetable_path")),
        adjusted_default_path,
    )
    adjusted_timetable_sheet_name = _str_or_default(
        analyze_cfg.get("adj_timetable_sheet_name", analyze_cfg.get("adjusted_timetable_sheet_name")),
        "Sheet1",
    )
    metrics_output_path = _path_or_default(analyze_cfg.get("metrics_output_path"), metrics_default_path)
    plot_output_path = _path_or_default(analyze_cfg.get("plot_output_path"), plot_default_path)

    delays = [
        DelayScenario(
            train_id=str(item["train_id"]).strip(),
            station=str(item["station"]).strip(),
            event_type=str(item["event_type"]).strip(),
            seconds=int(item["seconds"]),
        )
        for item in scenarios_cfg.get("delays", [])
    ]
    speed_limits = [
        SpeedLimitScenario(
            start_station=str(item["start_station"]).strip(),
            end_station=str(item["end_station"]).strip(),
            extra_seconds=int(item["extra_seconds"]),
            start_time=_parse_time_to_seconds(item["start_time"]),
            end_time=_parse_time_to_seconds(item["end_time"]),
        )
        for item in scenarios_cfg.get("speed_limits", [])
    ]
    interruptions = [
        InterruptionScenario(
            start_station=str(item["start_station"]).strip(),
            end_station=str(item["end_station"]).strip(),
            start_time=_parse_time_to_seconds(item["start_time"]),
            end_time=_parse_time_to_seconds(item["end_time"]),
        )
        for item in scenarios_cfg.get("interruptions", [])
    ]

    objective_mode_raw = str(_solve_value("objective_mode", "abs")).strip()
    cancellation_default = False
    objective_mode = objective_mode_raw
    if objective_mode_raw == "cal_delay_plus_cancel":
        # Compatibility alias: old mixed mode now maps to
        # abs objective + independent cancellation switch.
        objective_mode = "abs"
        cancellation_default = True

    return AppConfig(
        project=ProjectConfig(name=case_name, output_dir=output_dir),
        input=InputConfig(
            timetable_path=timetable_path,
            mileage_path=mileage_path,
            timetable_sheet_name=timetable_sheet_name,
            mileage_sheet_name=mileage_sheet_name,
        ),
        solver=SolverConfig(
            objective_delay_weight=float(_solve_value("objective_delay_weight", 1.0)),
            objective_mode=objective_mode,
            cancellation_enabled=_bool_or_default(
                _solve_value("cancellation_enabled", cancellation_default),
                cancellation_default,
            ),
            cancellation_penalty_weight=float(
                _solve_value("cancellation_penalty_weight", 1000.0)
            ),
            arr_arr_headway_seconds=int(_solve_value("arr_arr_headway_seconds", 180)),
            dep_dep_headway_seconds=int(_solve_value("dep_dep_headway_seconds", 180)),
            dwell_seconds_at_stops=int(_solve_value("dwell_seconds_at_stops", 120)),
            big_m=int(_solve_value("big_m", 100000)),
            tolerance_delay_seconds=int(
                _solve_value(
                    "cancellation_threshold_seconds",
                    _solve_value("tolerance_delay_seconds", 2 * 3600),
                )
            ),
        ),
        scenarios=ScenarioConfig(
            delays=delays,
            speed_limits=speed_limits,
            interruptions=interruptions,
        ),
        build=BuildConfig(lp_path=lp_default_path),
        solve=SolveConfig(lp_path=solve_lp_path, solution_path=solution_default_path),
        export_timetable=ExportTimetableConfig(
            solution_path=export_solution_path,
            timetable_path=adjusted_default_path,
        ),
        analyze=AnalyzeConfig(
            enable_metrics=bool(analyze_cfg.get("enable_metrics", False)),
            enable_plot=bool(analyze_cfg.get("enable_plot", False)),
            plot_grid=bool(analyze_cfg.get("plot_grid", False)),
            plot_title=str(analyze_cfg.get("plot_title", "Train Timetable")),
            plan_timetable_path=timetable_path,
            plan_timetable_sheet_name=timetable_sheet_name,
            adjusted_timetable_path=adjusted_timetable_path,
            adjusted_timetable_sheet_name=adjusted_timetable_sheet_name,
            metrics_output_path=metrics_output_path,
            plot_output_path=plot_output_path,
            plot_timetable_path=adjusted_timetable_path,
        ),
    )


def load_timetable(path: Path, sheet_name: str) -> RawTable:
    return _read_excel(path, sheet_name)


def load_mileage_table(path: Path, sheet_name: str) -> RawTable:
    return _read_excel(path, sheet_name)


