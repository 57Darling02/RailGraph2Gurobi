from __future__ import annotations

import argparse
import copy
import csv
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List

# Ensure project imports work when running as:
#   python scripts/batch_solution_analyze.py
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from main import cmd_analyze, cmd_export_timetable

def _require_yaml():
    try:
        import yaml
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Missing dependency: pyyaml") from exc
    return yaml

def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Batch run export-timetable + analyze for many .sol files and aggregate summary."
    )
    parser.add_argument("--solutions-root", default="tests/solutions", help="Root folder containing *.sol files.")
    parser.add_argument("--base-config", default="tests/test.yaml", help="Template config used as base.")
    parser.add_argument("--generated-config-root", default="tests/generated_configs", help="Generated yaml output root.")
    parser.add_argument("--output-root", default="outputs/solutions_batch", help="Per-solution pipeline output root.")
    parser.add_argument("--summary-csv", default="outputs/solutions_batch/summary.csv", help="Aggregated CSV path.")
    parser.add_argument("--summary-json", default="outputs/solutions_batch/summary.json", help="Aggregated JSON path.")
    parser.add_argument("--limit", type=int, default=0, help="Only process first N sol files (0 = all).")
    parser.add_argument(
        "--enable-plot",
        action="store_true",
        help="Enable plotting in analyze step. Default false to avoid plot failures on empty timetables.",
    )
    return parser.parse_args()


def _resolve_path(path_text: str) -> Path:
    path = Path(path_text)
    if path.is_absolute():
        return path
    return (REPO_ROOT / path).resolve()


def _sanitize_case_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-") or "case"


def _read_metrics_summary(metrics_path: Path) -> Dict[str, object]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Missing dependency: openpyxl") from exc

    workbook = load_workbook(metrics_path, read_only=True, data_only=True)
    if "summary" not in workbook.sheetnames:
        raise ValueError(f"Sheet 'summary' not found in metrics file: {metrics_path}")
    sheet = workbook["summary"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    return {str(h): v for h, v in zip(headers, values) if h is not None}

def _build_case_config(
    base_payload: Dict[str, object],
    case_name: str,
    output_dir: Path,
    enable_plot: bool,
) -> Dict[str, object]:
    payload = copy.deepcopy(base_payload)

    project = dict(payload.get("project", {}))
    legacy_input = dict(payload.get("input", {}))
    required_project_defaults = {
        "timetable_path": "",
        "mileage_path": "",
        "timetable_sheet_name": "Sheet1",
        "mileage_sheet_name": "Sheet1",
    }
    for key, default_value in required_project_defaults.items():
        current_value = project.get(key)
        current_text = "" if current_value is None else str(current_value).strip()
        if current_text != "":
            continue
        legacy_value = legacy_input.get(key, default_value)
        project[key] = default_value if legacy_value is None else legacy_value

    project["name"] = case_name
    project["output_dir"] = str(output_dir).replace("\\", "/")
    payload["project"] = project

    build = dict(payload.get("build", {}))
    build["scenarios"] = {
        "delays": [],
        "speed_limits": [],
        "interruptions": [],
    }
    payload["build"] = build

    solve = dict(payload.get("solve", {}))
    if not solve:
        solve = dict(payload.get("solver", {}))
    solve["lp_path"] = ""
    payload["solve"] = solve

    export_cfg = dict(payload.get("export-timetable", payload.get("export_timetable", {})))
    export_cfg["sol_path"] = ""
    payload["export-timetable"] = export_cfg
    payload.pop("export_timetable", None)

    analyze = dict(payload.get("analyze", {}))
    analyze["enable_metrics"] = True
    analyze["enable_plot"] = bool(enable_plot)
    analyze.setdefault("plot_grid", False)
    analyze.setdefault("plot_title", "Train Timetable")
    analyze.setdefault("adj_timetable_path", "")
    analyze.setdefault("adj_timetable_sheet_name", "Sheet1")
    payload["analyze"] = analyze

    payload.pop("scenarios", None)
    return payload


def main() -> None:
    args = _parse_args()
    yaml = _require_yaml()
    solutions_root = _resolve_path(args.solutions_root)
    base_config_path = _resolve_path(args.base_config)
    generated_config_root = _resolve_path(args.generated_config_root)
    output_root = _resolve_path(args.output_root)
    summary_csv = _resolve_path(args.summary_csv)
    summary_json = _resolve_path(args.summary_json)

    if not solutions_root.exists():
        raise FileNotFoundError(f"solutions root not found: {solutions_root}")
    if not base_config_path.exists():
        raise FileNotFoundError(f"base config not found: {base_config_path}")

    with base_config_path.open("r", encoding="utf-8") as file:
        base_payload = yaml.safe_load(file) or {}

    sol_files = sorted(solutions_root.rglob("*.sol"))
    if args.limit > 0:
        sol_files = sol_files[: args.limit]

    generated_config_root.mkdir(parents=True, exist_ok=True)
    output_root.mkdir(parents=True, exist_ok=True)
    summary_csv.parent.mkdir(parents=True, exist_ok=True)
    summary_json.parent.mkdir(parents=True, exist_ok=True)

    records: List[Dict[str, object]] = []

    for idx, sol_path in enumerate(sol_files, start=1):
        rel = sol_path.relative_to(solutions_root)
        rel_no_ext = rel.with_suffix("")
        case_key = _sanitize_case_name(str(rel_no_ext).replace("\\", "/"))
        case_output_dir = output_root / rel_no_ext
        case_config_path = generated_config_root / rel_no_ext.parent / f"{rel_no_ext.name}.yaml"

        case_config_path.parent.mkdir(parents=True, exist_ok=True)
        case_output_dir.mkdir(parents=True, exist_ok=True)

        case_payload = _build_case_config(
            base_payload=base_payload,
            case_name=case_key,
            output_dir=case_output_dir,
            enable_plot=args.enable_plot,
        )
        with case_config_path.open("w", encoding="utf-8") as file:
            yaml.safe_dump(case_payload, file, allow_unicode=True, sort_keys=False)

        expected_solution_path = case_output_dir / f"{case_key}.sol"
        shutil.copy2(sol_path, expected_solution_path)

        record: Dict[str, object] = {
            "index": idx,
            "solution_file": str(sol_path).replace("\\", "/"),
            "generated_config": str(case_config_path).replace("\\", "/"),
            "output_dir": str(case_output_dir).replace("\\", "/"),
            "status": "ok",
            "error": "",
            "late_train_count": "",
            "late_train_ids": "",
            "end_late_train_count": "",
            "canceled_train_count": "",
            "canceled_train_ids": "",
            "total_late_sec": "",
            "total_deviation_sec": "",
            "metrics_file": "",
            "adjusted_timetable_file": str(case_output_dir / "adjusted_timetable.xlsx").replace("\\", "/"),
            "plot_file": str(case_output_dir / "timetable_plot.png").replace("\\", "/"),
        }

        try:
            cmd_export_timetable(case_config_path)
            cmd_analyze(case_config_path)

            metrics_path = case_output_dir / "analysis_metrics.xlsx"
            record["metrics_file"] = str(metrics_path).replace("\\", "/")
            if metrics_path.exists():
                summary = _read_metrics_summary(metrics_path)
                record["late_train_count"] = summary.get("late_train_count", "")
                record["late_train_ids"] = summary.get("late_train_ids", "")
                record["end_late_train_count"] = summary.get("end_late_train_count", "")
                record["canceled_train_count"] = summary.get("canceled_train_count", "")
                record["canceled_train_ids"] = summary.get("canceled_train_ids", "")
                record["total_late_sec"] = summary.get("total_late_sec", "")
                record["total_deviation_sec"] = summary.get("total_deviation_sec", "")
            else:
                record["status"] = "metrics_missing"
        except Exception as exc:  # pragma: no cover
            record["status"] = "failed"
            record["error"] = str(exc)

        records.append(record)

    fieldnames = list(records[0].keys()) if records else [
        "index",
        "solution_file",
        "generated_config",
        "output_dir",
        "status",
        "error",
    ]
    with summary_csv.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    status_counts: Dict[str, int] = {}
    late_sum = 0
    for item in records:
        status = str(item.get("status", "unknown"))
        status_counts[status] = status_counts.get(status, 0) + 1
        try:
            late_sum += int(item.get("late_train_count") or 0)
        except ValueError:
            pass

    summary_payload = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "solutions_root": str(solutions_root).replace("\\", "/"),
        "base_config": str(base_config_path).replace("\\", "/"),
        "generated_config_root": str(generated_config_root).replace("\\", "/"),
        "output_root": str(output_root).replace("\\", "/"),
        "total_solutions": len(records),
        "status_counts": status_counts,
        "sum_late_train_count": late_sum,
        "summary_csv": str(summary_csv).replace("\\", "/"),
    }
    with summary_json.open("w", encoding="utf-8") as file:
        json.dump(summary_payload, file, ensure_ascii=False, indent=2)

    print(f"Processed solutions: {len(records)}")
    print(f"Status counts: {status_counts}")
    print(f"Sum late_train_count: {late_sum}")
    print(f"Summary CSV: {summary_csv}")
    print(f"Summary JSON: {summary_json}")


if __name__ == "__main__":
    main()



